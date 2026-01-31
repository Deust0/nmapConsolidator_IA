"""
Módulo para generar archivos Excel con todas las hojas y formatos
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime
from config import COLORES_SEVERIDAD
from nmap_commands import obtener_scripts_para_puerto, generar_comando_nmap_inteligente
from ai_analyzer import analizar_resultados_completos, obtener_resumen_vulnerabilidades


def aplicar_formato_encabezado(ws, fila=1):
    """Aplica formato estándar a los encabezados"""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[fila]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def configurar_impresion(ws):
    """Configura la impresión de la hoja para ser profesional y legible"""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = False
    ws.print_rows = f"1:1"


def crear_hoja_resultados(wb, resultados):
    """Crea hoja de resultados consolidados - AÑADIDA COLUMNA HOSTNAMES"""
    ws = wb.active
    ws.title = "Resultados Consolidados"
    
    # Columnas reorganizadas: Estado Auditoría al lado de Severidad, eliminadas OS Type, Método, Confianza, Scripts
    # Agregadas columnas de CVSS 3.1
    encabezados = [
        "ID", "Archivo Origen", "IP", "Hostnames/Dominios", "Puerto", "Protocolo", "Estado Puerto",
        "Servicio", "Versión", "Vulnerabilidades Detectadas", "Severidad", "Estado Auditoría",
        "AV", "AC", "PR", "UI", "S", "C", "I", "A", "CVSS 3.1", "Riesgo Calculado", "Resumen", "Impacto", "Recomendación",
        "Observaciones", "Fecha Detección", "CPEs"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    # Data Validation para Severidad (Col K)
    dv_severidad = DataValidation(type="list", formula1='"Crítica,Alta,Media,Baja,Revisado - No vulnerable,Pendiente"', allow_blank=True)
    dv_severidad.error = "Selecciona: Crítica, Alta, Media, Baja, Revisado - No vulnerable, Pendiente"
    dv_severidad.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_severidad)
    
    # Data Validation para Estado Auditoría (Col L, al lado de Severidad)
    dv_estado = DataValidation(type="list", formula1='"Pendiente,En Proceso,Finalizado,Aceptado"', allow_blank=True)
    dv_estado.error = "Selecciona: Pendiente, En Proceso, Finalizado, Aceptado"
    dv_estado.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_estado)
    
    # Data Validation para componentes CVSS
    dv_av = DataValidation(type="list", formula1='"Network,Adjacent,Local,Physical"', allow_blank=True)
    dv_ac = DataValidation(type="list", formula1='"Low,High"', allow_blank=True)
    dv_pr = DataValidation(type="list", formula1='"None,Low,High"', allow_blank=True)
    dv_ui = DataValidation(type="list", formula1='"None,Required"', allow_blank=True)
    dv_s = DataValidation(type="list", formula1='"Unchanged,Changed"', allow_blank=True)
    dv_cia = DataValidation(type="list", formula1='"None,Low,High"', allow_blank=True)
    
    ws.add_data_validation(dv_av)
    ws.add_data_validation(dv_ac)
    ws.add_data_validation(dv_pr)
    ws.add_data_validation(dv_ui)
    ws.add_data_validation(dv_s)
    ws.add_data_validation(dv_cia)
    
    for idx, row_data in enumerate(resultados, 2):
        # Referencias de columnas:
        # K: Severidad, L: Estado Auditoría
        # M: AV, N: AC, O: PR, P: UI, Q: S, R: C, S: I, T: A
        # U: CVSS 3.1 (calculado), V: Riesgo Calculado
        # W: Resumen, X: Impacto, Y: Recomendación, Z: Observaciones
        # AA: Fecha Detección, AB: CPEs
        
        # Fórmula CVSS 3.1 simplificada (requiere implementación completa en producción)
        # Si Severidad es "Revisado - No vulnerable" o contiene "Falso Positivo", CVSS = 0
        cvss_formula = f'=IF(OR(K{idx}="Revisado - No vulnerable",K{idx}="Falso Positivo"),0,IF(AND(M{idx}<>"",N{idx}<>"",O{idx}<>"",P{idx}<>"",Q{idx}<>"",R{idx}<>"",S{idx}<>"",T{idx}<>""),ROUND((IF(M{idx}="Network",0.85,IF(M{idx}="Adjacent",0.62,IF(M{idx}="Local",0.55,0.2)))*IF(N{idx}="Low",0.77,0.44)*IF(O{idx}="None",0.85,IF(O{idx}="Low",IF(Q{idx}="Unchanged",0.62,0.68),IF(Q{idx}="Unchanged",0.27,0.5)))*IF(P{idx}="None",0.85,0.62)+IF(R{idx}="High",0.56,IF(R{idx}="Low",0.22,0))*IF(S{idx}="High",0.56,IF(S{idx}="Low",0.22,0))*IF(T{idx}="High",0.56,IF(T{idx}="Low",0.22,0)))*10,1),""))'
        
        row = [
            idx - 1,
            row_data.get("archivo_origen", ""),
            row_data.get("ip", ""),
            row_data.get("hostnames", ""),
            row_data.get("puerto", ""),
            row_data.get("protocolo", ""),
            row_data.get("estado", "open"),
            row_data.get("servicio", ""),
            row_data.get("version", ""),
            "",  # Vulnerabilidades Detectadas (J)
            "Pendiente",  # Severidad (K)
            "Pendiente",  # Estado Auditoría (L)
            "",  # AV - Attack Vector (M)
            "",  # AC - Attack Complexity (N)
            "",  # PR - Privileges Required (O)
            "",  # UI - User Interaction (P)
            "",  # S - Scope (Q)
            "",  # C - Confidentiality (R)
            "",  # I - Integrity (S)
            "",  # A - Availability (T)
            cvss_formula,  # CVSS 3.1 (U) - calculado
            f"=IF(U{idx}=0,0,IF(U{idx}=\"\",\"\",U{idx}*IF(K{idx}=\"Crítica\",4,IF(K{idx}=\"Alta\",3,IF(K{idx}=\"Media\",2,IF(K{idx}=\"Baja\",1,0)))))",  # Riesgo Calculado (V)
            "",  # Resumen (W)
            "",  # Impacto (X)
            "",  # Recomendación (Y)
            "",  # Observaciones (Z)
            datetime.now().strftime("%Y-%m-%d"),  # Fecha Detección (AA)
            row_data.get("cpes", "")  # CPEs (AB)
        ]
        ws.append(row)
        
        # Aplicar validaciones
        dv_severidad.add(f"K{idx}")
        dv_estado.add(f"L{idx}")
        dv_av.add(f"M{idx}")
        dv_ac.add(f"N{idx}")
        dv_pr.add(f"O{idx}")
        dv_ui.add(f"P{idx}")
        dv_s.add(f"Q{idx}")
        dv_cia.add(f"R{idx}")  # Confidentiality
        dv_cia.add(f"S{idx}")  # Integrity
        dv_cia.add(f"T{idx}")  # Availability
        
        severidad_celda = ws[f"K{idx}"]
        severidad_celda.fill = PatternFill(start_color=COLORES_SEVERIDAD["Pendiente"], 
                                           end_color=COLORES_SEVERIDAD["Pendiente"], 
                                           fill_type="solid")
        
        estado_celda = ws[f"L{idx}"]
        estado_celda.fill = PatternFill(start_color="CCCCCC", 
                                       end_color="CCCCCC", 
                                       fill_type="solid")
    
    # Ajustar anchos de columnas (28 columnas ahora)
    columnas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']
    anchos = [5, 12, 13, 25, 8, 10, 12, 12, 10, 20, 12, 12, 8, 8, 8, 8, 8, 8, 8, 8, 10, 12, 20, 20, 20, 20, 12, 25]
    
    for col, ancho in zip(columnas, anchos):
        ws.column_dimensions[col].width = ancho
    
    configurar_impresion(ws)


def crear_hoja_dashboard_vulnerabilidades(wb, resultados):
    """Crea dashboard con gráficos coloreados - ACTUALIZADO REFERENCIAS DE COLUMNAS"""
    ws = wb.create_sheet(title="Dashboard Vulnerabilidades", index=1)
    
    header_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws['A1'] = "DASHBOARD DE VULNERABILIDADES"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # Estado de Revisión
    ws['A3'] = "ESTADO DE REVISIÓN"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:B3')
    
    ws['A4'] = "Estado"
    ws['B4'] = "Cantidad"
    for cell in [ws['A4'], ws['B4']]:
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
    
    estados = ["Pendiente", "En Proceso", "Finalizado", "Aceptado"]
    max_row = len(resultados) + 100
    
    # Actualizado referencia: Estado Auditoría ahora es L
    for idx, estado in enumerate(estados, 5):
        ws[f'A{idx}'] = estado
        ws[f'B{idx}'] = f"=COUNTIF('Resultados Consolidados'!$L$2:$L${max_row},A{idx})"
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    
    # Severidad
    ws['D3'] = "VULNERABILIDADES POR SEVERIDAD"
    ws['D3'].font = Font(bold=True, size=11)
    ws['D3'].fill = header_fill
    ws.merge_cells('D3:E3')
    
    ws['D4'] = "Severidad"
    ws['E4'] = "Cantidad"
    for cell in [ws['D4'], ws['E4']]:
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Actualizado referencia: Severidad ahora es K
    severidades = ["Crítica", "Alta", "Media", "Baja", "Revisado - No vulnerable", "Pendiente"]
    for idx, severidad in enumerate(severidades, 5):
        celda_sev = ws[f'D{idx}']
        celda_sev.value = severidad
        celda_sev.fill = PatternFill(start_color=COLORES_SEVERIDAD[severidad], 
                                      end_color=COLORES_SEVERIDAD[severidad], 
                                      fill_type="solid")
        celda_sev.font = Font(bold=True, color="FFFFFF" if severidad in ["Crítica", "Alta"] else "000000")
        
        ws[f'E{idx}'] = f"=COUNTIF('Resultados Consolidados'!$K$2:$K${max_row},D{idx})"
    
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    
    # Gráficos
    pie1 = PieChart()
    pie1.title = "Distribución por Estado de Revisión"
    pie1.style = 10
    labels1 = Reference(ws, min_col=1, min_row=5, max_row=8)
    data1 = Reference(ws, min_col=2, min_row=4, max_row=8)
    pie1.add_data(data1, titles_from_data=True)
    pie1.set_categories(labels1)
    ws.add_chart(pie1, "A10")
    
    pie2 = PieChart()
    pie2.title = "Distribución por Severidad"
    pie2.style = 11
    labels2 = Reference(ws, min_col=4, min_row=5, max_row=10)
    data2 = Reference(ws, min_col=5, min_row=4, max_row=10)
    pie2.add_data(data2, titles_from_data=True)
    pie2.set_categories(labels2)
    
    colores_hex = [COLORES_SEVERIDAD[s] for s in severidades]
    for idx, color_hex in enumerate(colores_hex):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color_hex
        pie2.series[0].data_points.append(pt)
    
    ws.add_chart(pie2, "D10")
    configurar_impresion(ws)


def crear_hoja_resumen_ips(wb, resultados):
    """Crea hoja de resumen por IP - ACTUALIZADO REFERENCIAS DE COLUMNAS"""
    ws = wb.create_sheet(title="Resumen por IP")
    ws.append(["IP", "Dominios Detectados", "Total Puertos", "Servicios Principales", "Riesgo Máximo", "Riesgo Promedio"])
    aplicar_formato_encabezado(ws)
    
    ips_dict = {}
    for row_data in resultados:
        ip = row_data.get("ip", "")
        if ip not in ips_dict:
            ips_dict[ip] = []
        ips_dict[ip].append(row_data)
    
    max_row_res = len(resultados) + 100
    
    for idx, (ip, datos) in enumerate(sorted(ips_dict.items()), 2):
        servicios = list(set([d.get("servicio", "Unknown") for d in datos if d.get("servicio")]))
        servicios_str = ", ".join(servicios[:5])
        
        # Obtener hostnames (usando el primero que encuentre no vacío)
        hostnames = ""
        for d in datos:
            if d.get("hostnames"):
                hostnames = d.get("hostnames")
                break

        # Referencias: IP (Col C), Riesgo Calculado (Col N)
        ws.append([
            ip,
            hostnames,
            len(datos),
            servicios_str,
            f"=IFERROR(MAX(IF('Resultados Consolidados'!$C$2:$C${max_row_res}=A{idx},'Resultados Consolidados'!$N$2:$N${max_row_res})),\"\")",
            f"=IFERROR(AVERAGE(IF('Resultados Consolidados'!$C$2:$C${max_row_res}=A{idx},'Resultados Consolidados'!$N$2:$N${max_row_res})),\"\")"
        ])
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    configurar_impresion(ws)


def crear_hoja_seguimiento_vulnerabilidades(wb, resultados):
    """Crea hoja de seguimiento centralizado - Estado Auditoría al lado de Severidad"""
    ws = wb.create_sheet(title="Seguimiento Vulnerabilidades")
    
    encabezados = [
        "Hallazgo ID", "IP", "Hostname", "Puerto", "Servicio", "Vulnerabilidad",
        "Severidad", "Estado Auditoría", "CVSS", "Estado Remediación", "Fecha Detección",
        "Fecha Resolución", "Responsable", "Observaciones"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    # Data Validation para Severidad (Col G)
    dv_severidad = DataValidation(type="list", 
                                   formula1='"Crítica,Alta,Media,Baja,Revisado - No vulnerable,Pendiente"', 
                                   allow_blank=True)
    dv_severidad.error = "Selecciona: Crítica, Alta, Media, Baja, Revisado - No vulnerable, Pendiente"
    dv_severidad.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_severidad)
    
    # Data Validation para Estado Auditoría (Col H, al lado de Severidad)
    dv_estado_aud = DataValidation(type="list", 
                                    formula1='"Pendiente,En Proceso,Finalizado,Aceptado"', 
                                    allow_blank=True)
    dv_estado_aud.error = "Selecciona: Pendiente, En Proceso, Finalizado, Aceptado"
    dv_estado_aud.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_estado_aud)
    
    # Data Validation para Estado Remediación (Col J)
    dv_estado_rem = DataValidation(type="list", 
                                    formula1='"Identificada,En Remediación,Remediada,Aceptada"', 
                                    allow_blank=True)
    ws.add_data_validation(dv_estado_rem)
    
    for idx, row_data in enumerate(resultados, 2):
        ws.append([
            f"HALL-{idx-1:04d}",
            row_data.get("ip", ""),
            row_data.get("hostnames", ""),
            row_data.get("puerto", ""),
            row_data.get("servicio", ""),
            "",
            "Pendiente",  # Severidad (G)
            "Pendiente",  # Estado Auditoría (H) - ahora al lado de Severidad
            "",  # CVSS (I)
            "Identificada",  # Estado Remediación (J)
            datetime.now().strftime("%Y-%m-%d"),
            "",
            "",
            ""
        ])
        dv_severidad.add(f"G{idx}")
        dv_estado_aud.add(f"H{idx}")
        dv_estado_rem.add(f"J{idx}")
        
        # Aplicar color a Severidad
        severidad_celda = ws[f"G{idx}"]
        severidad_celda.fill = PatternFill(start_color=COLORES_SEVERIDAD["Pendiente"], 
                                           end_color=COLORES_SEVERIDAD["Pendiente"], 
                                           fill_type="solid")
        
        # Aplicar color a Estado Auditoría
        estado_aud_celda = ws[f"H{idx}"]
        estado_aud_celda.fill = PatternFill(start_color="CCCCCC", 
                                           end_color="CCCCCC", 
                                           fill_type="solid")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 15
    
    configurar_impresion(ws)


def crear_hoja_matriz_riesgos(wb, resultados):
    """Crea matriz de riesgos - ACTUALIZADO REFERENCIAS A SEGUIMIENTO"""
    ws = wb.create_sheet(title="Matriz de Riesgos")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = "MATRIZ DE RIESGOS Y GESTIÓN DE VULNERABILIDADES"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:F1')
    
    ws['A3'] = "MATRIZ: SEVERIDAD x ESTADO REMEDIACIÓN"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:E3')
    
    ws['A4'] = "Severidad"
    ws['B4'] = "Identificada"
    ws['C4'] = "En Remediación"
    ws['D4'] = "Remediada"
    ws['E4'] = "Total"
    
    for cell in [ws['A4'], ws['B4'], ws['C4'], ws['D4'], ws['E4']]:
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = header_font
    
    max_row_vuln = len(resultados) + 100
    severidades = ["Crítica", "Alta", "Media", "Baja"]
    
    # En Seguimiento: Severidad es Col G (7), Estado Remediación es Col J (10)
    for idx, severidad in enumerate(severidades, 5):
        celda = ws[f'A{idx}']
        celda.value = severidad
        celda.fill = PatternFill(start_color=COLORES_SEVERIDAD[severidad], 
                                 end_color=COLORES_SEVERIDAD[severidad], 
                                 fill_type="solid")
        celda.font = Font(bold=True, color="FFFFFF" if severidad in ["Crítica", "Alta"] else "000000")
        
        ws[f'B{idx}'] = f"=COUNTIFS('Seguimiento Vulnerabilidades'!$G$2:$G${max_row_vuln},A{idx},'Seguimiento Vulnerabilidades'!$J$2:$J${max_row_vuln},\"Identificada\")"
        ws[f'C{idx}'] = f"=COUNTIFS('Seguimiento Vulnerabilidades'!$G$2:$G${max_row_vuln},A{idx},'Seguimiento Vulnerabilidades'!$J$2:$J${max_row_vuln},\"En Remediación\")"
        ws[f'D{idx}'] = f"=COUNTIFS('Seguimiento Vulnerabilidades'!$G$2:$G${max_row_vuln},A{idx},'Seguimiento Vulnerabilidades'!$J$2:$J${max_row_vuln},\"Remediada\")"
        ws[f'E{idx}'] = f"=B{idx}+C{idx}+D{idx}"
    
    ws['A9'] = "TOTAL"
    ws['A9'].font = Font(bold=True)
    ws['A9'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws['B9'] = "=SUM(B5:B8)"
    ws['C9'] = "=SUM(C5:C8)"
    ws['D9'] = "=SUM(D5:D8)"
    ws['E9'] = "=SUM(E5:E8)"
    
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Matriz de Riesgos por Severidad y Estado"
    data = Reference(ws, min_col=2, min_row=4, max_col=4, max_row=8)
    cats = Reference(ws, min_col=1, min_row=5, max_row=8)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "A11")
    configurar_impresion(ws)


def crear_hoja_analisis_ip(wb, resultados):
    """Crea tabla dinámica de análisis por IP - ACTUALIZADO REFERENCIAS"""
    ws = wb.create_sheet(title="Análisis por IP")
    
    encabezados = [
        "IP", "Hostname", "Total Puertos", "Servicios Únicos", "Puertos Críticos",
        "Puertos Altos", "Puertos Medios", "Puertos Bajos",
        "Pendientes", "En Proceso", "Finalizados", "No Vulnerables"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    ips_dict = {}
    for row_data in resultados:
        ip = row_data.get("ip", "")
        if ip not in ips_dict:
            ips_dict[ip] = []
        ips_dict[ip].append(row_data)
    
    max_row_res = len(resultados) + 100
    
    for idx, (ip, datos) in enumerate(sorted(ips_dict.items()), 2):
        servicios = list(set([d.get("servicio", "") for d in datos if d.get("servicio")]))
        hostnames = ""
        for d in datos:
            if d.get("hostnames"):
                hostnames = d.get("hostnames")
                break
        
        # Referencias en Consolidados:
        # IP = C, Severidad = K, Estado Auditoría = L
        ws.append([
            ip, hostnames, len(datos), len(servicios),
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Crítica\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Alta\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Media\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Baja\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$L$2:$L${max_row_res},\"Pendiente\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$L$2:$L${max_row_res},\"En Proceso\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$L$2:$L${max_row_res},\"Finalizado\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Revisado - No vulnerable\")"
        ])
    
    for col in range(1, 13):
        ws.column_dimensions[chr(64 + col)].width = 14
    ws.column_dimensions['B'].width = 25
    configurar_impresion(ws)


def crear_hoja_comandos_mejorada(wb, resultados):
    """Crea hoja de comandos Nmap inteligentes"""
    ws = wb.create_sheet(title="Comandos Nmap Inteligentes")
    
    encabezados = [
        "IP", "Hostname", "Puerto", "Servicio", "Scripts Disponibles",
        "Comando Nmap (NO INTRUSIVO)", "Archivo Output", "Estado"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    ips_puertos = {}
    for row_data in resultados:
        ip = row_data.get("ip", "")
        hostname = row_data.get("hostnames", "")
        puerto = row_data.get("puerto", "")
        servicio = row_data.get("servicio", "")
        version = row_data.get("version", "")
        
        if ip not in ips_puertos:
            ips_puertos[ip] = []
        ips_puertos[ip].append((puerto, servicio, version, hostname))
    
    for ip, puertos_lista in sorted(ips_puertos.items()):
        for puerto, servicio, version, hostname in sorted(puertos_lista):
            scripts = obtener_scripts_para_puerto(puerto, servicio)
            scripts_str = ", ".join(scripts)
            
            comando = generar_comando_nmap_inteligente(ip, puerto, servicio, version)
            archivo_output = f"nmap_{ip}_{puerto}.txt"
            
            ws.append([
                ip, hostname, puerto, servicio, scripts_str,
                comando, archivo_output, "Pendiente"
            ])
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 100
    configurar_impresion(ws)


def crear_hoja_info_escaneos(wb, archivos_procesados):
    """Crea hoja con información de escaneos"""
    ws = wb.create_sheet(title="Info Escaneos")
    ws.append(["Archivo Procesado", "Ruta Completa", "Fecha Procesamiento"])
    aplicar_formato_encabezado(ws)
    for archivo in archivos_procesados:
        ws.append([os.path.basename(archivo), archivo, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    configurar_impresion(ws)


def crear_hoja_instrucciones(wb):
    """Crea hoja con instrucciones de uso"""
    ws = wb.create_sheet(title="Instrucciones")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws['A1'] = "GUÍA DE USO - GESTOR DE VULNERABILIDADES v9"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')
    
    instrucciones = [
        "", "NOVEDADES v9.1:", "=================", 
        "1. SOPORTE PARA ARCHIVOS .NMAP",
        "   - Procesa archivos .nmap además de .gnmap y .xml",
        "   - Extrae dominios COMPLETOS (con TLD) de líneas 'Nmap scan report for'",
        "   - Identifica múltiples subdominios por IP",
        "   - NO usa hostnames de .gnmap (están incompletos, sin TLD)",
        "",
        "NOVEDADES v9:", "=================", 
        "1. COLUMNA HOSTNAMES/DOMINIOS",
        "   - Nueva columna en 'Resultados Consolidados' y otras hojas",
        "   - Extrae nombres DNS de archivos XML (<hostnames>)",
        "   - Extrae dominios completos de archivos NMAP (líneas de reporte)",
        "   - NO extrae hostnames de archivos GNMAP (incompletos)",
        "",
        "HOJAS DISPONIBLES:", "=================",
        "1. RESULTADOS CONSOLIDADOS",
        "   - Columna Hostnames añadida después de IP",
        "   - Riesgo Calculado ajustado a nuevas columnas",
        "",
        "2. DASHBOARD VULNERABILIDADES",
        "   - Gráficos actualizados dinámicamente",
        "",
        "3. RESUMEN POR IP",
        "   - Incluye columna de dominios asociados a la IP",
        "",
        "4. COMANDOS NMAP INTELIGENTES",
        "   - Incluye referencia del hostname",
        ""
    ]
    
    for idx, instruccion in enumerate(instrucciones, 3):
        ws[f'A{idx}'] = instruccion
    
    ws.column_dimensions['A'].width = 100
    configurar_impresion(ws)


def crear_hoja_vulnerabilidades_ia(wb, vulnerabilidades_ia):
    """Crea hoja con vulnerabilidades detectadas por IA"""
    if not vulnerabilidades_ia:
        return
    
    ws = wb.create_sheet(title="Vulnerabilidades IA")
    
    encabezados = [
        "ID", "IP", "Puerto", "Servicio", "Tipo", "Vector Ataque", "Severidad", "CVSS 3.1", 
        "Criticidad Ajustada", "Descripción", "Recomendación", "Estado"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    # Data Validation para Vector de Ataque
    dv_vector = DataValidation(type="list", formula1='"Internet,Adyacente,Local,Físico"', allow_blank=True)
    dv_vector.error = "Selecciona: Internet, Adyacente, Local, Físico"
    dv_vector.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_vector)
    
    # Data Validation para Severidad
    dv_severidad = DataValidation(type="list", formula1='"Crítica,Alta,Media,Baja"', allow_blank=True)
    dv_severidad.error = "Selecciona: Crítica, Alta, Media, Baja"
    dv_severidad.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_severidad)
    
    # Data Validation para Estado
    dv_estado = DataValidation(type="list", formula1='"Pendiente,Verificado,Falso Positivo,Confirmado"', allow_blank=True)
    dv_estado.error = "Selecciona: Pendiente, Verificado, Falso Positivo, Confirmado"
    dv_estado.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_estado)
    
    for vuln in vulnerabilidades_ia:
        idx = ws.max_row + 1
        
        # Fórmula CVSS: Si Estado es "Falso Positivo", CVSS = 0
        # Si no, calcular basado en Vector de Ataque y Severidad
        cvss_formula = f'=IF(L{idx}="Falso Positivo",0,IF(AND(F{idx}<>"",G{idx}<>""),IF(F{idx}="Internet",IF(G{idx}="Crítica",9.0,IF(G{idx}="Alta",7.5,IF(G{idx}="Media",5.0,2.5))),IF(F{idx}="Adyacente",IF(G{idx}="Crítica",8.5,IF(G{idx}="Alta",7.0,IF(G{idx}="Media",4.5,2.0))),IF(F{idx}="Local",IF(G{idx}="Crítica",7.5,IF(G{idx}="Alta",6.5,IF(G{idx}="Media",4.0,1.5))),IF(G{idx}="Crítica",6.0,IF(G{idx}="Alta",5.0,IF(G{idx}="Media",3.5,1.0))))))),""))'
        
        # Fórmula Criticidad Ajustada: Basada en Vector de Ataque y Severidad
        criticidad_formula = f'=IF(L{idx}="Falso Positivo","Falso Positivo",IF(AND(F{idx}<>"",G{idx}<>""),IF(F{idx}="Internet",G{idx},IF(F{idx}="Adyacente",IF(G{idx}="Crítica","Crítica",IF(G{idx}="Alta","Alta","Media")),IF(F{idx}="Local",IF(G{idx}="Crítica","Alta",IF(G{idx}="Alta","Media","Baja")),IF(G{idx}="Crítica","Media",IF(G{idx}="Alta","Media","Baja"))))),""))'
        
        row = [
            vuln.get("id", ""),
            vuln.get("ip", ""),
            vuln.get("puerto", ""),
            vuln.get("servicio", ""),
            vuln.get("tipo", ""),
            "",  # Vector Ataque (F) - se llena manualmente
            vuln.get("severidad", "Pendiente"),  # Severidad (G)
            cvss_formula,  # CVSS 3.1 (H) - calculado
            criticidad_formula,  # Criticidad Ajustada (I) - calculada
            vuln.get("descripcion", ""),
            vuln.get("recomendacion", ""),
            vuln.get("estado", "Pendiente")
        ]
        ws.append(row)
        
        idx = ws.max_row
        dv_vector.add(f"F{idx}")
        dv_severidad.add(f"G{idx}")
        dv_estado.add(f"L{idx}")
        
        # Aplicar color según severidad
        severidad = vuln.get("severidad", "Pendiente")
        color_hex = COLORES_SEVERIDAD.get(severidad, COLORES_SEVERIDAD["Pendiente"])
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        ws[f"G{idx}"].fill = fill
    
    # Ajustar anchos de columna
    columnas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    anchos = [5, 13, 8, 12, 25, 12, 12, 10, 15, 40, 40, 15]
    
    for col, ancho in zip(columnas, anchos):
        ws.column_dimensions[col].width = ancho
    
    configurar_impresion(ws)
