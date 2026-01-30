###
# Script profesional de gestión de vulnerabilidades - VERSIÓN 9 (CON DOMINIOS)
# Características: Impresión profesional, Management de riesgos, Detección SSL/TLS
# NUEVAS v9: Columna "Hostnames/Dominios" para identificar activos por nombre DNS
# NUEVAS v9.1: Soporte para archivos .nmap además de .gnmap y .xml
###

import re
import os
import glob
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime


# Definir colores por severidad (RGB)
COLORES_SEVERIDAD = {
    "Crítica": "FF0000",                  # Rojo
    "Alta": "FF6600",                     # Naranja
    "Media": "FFD700",                    # Amarillo
    "Baja": "00B050",                     # Verde
    "Revisado - No vulnerable": "92D050", # Verde claro
    "Pendiente": "CCCCCC"                 # Gris
}

# Puertos típicos para escaneo SSL/TLS
PUERTOS_SSL_RECOMENDADOS = {
    "443", "8443", "9443",          # HTTPS variantes
    "993", "995",                   # IMAPS, POP3S
    "465", "587", "25",            # SMTPS / STARTTLS
    "110", "143",                  # POP3/IMAP STARTTLS
    "21", "990",                   # FTP/FTPS
    "389", "636",                  # LDAP/LDAPS
    "3389", "5986"                 # RDP TLS / WinRM
}

# Mapeo de puertos a scripts Nmap NO INTRUSIVOS
SCRIPTS_POR_PUERTO = {
    "21": {"servicio": "FTP", "scripts": ["ftp-anon", "ftp-bounce", "ftp-vsftpd-backdoor"]},
    "22": {"servicio": "SSH", "scripts": ["ssh-hostkey", "ssh-auth-methods", "ssh-brute"]},
    "25": {"servicio": "SMTP", "scripts": ["smtp-commands", "smtp-enum-users", "smtp-open-relay"]},
    "53": {"servicio": "DNS", "scripts": ["dns-brute", "dns-check-zone", "dns-nsid", "dns-recursion"]},
    "80": {"servicio": "HTTP", "scripts": ["http-title", "http-headers", "http-methods", "http-robots.txt", "http-git", "http-svn-enum", "http-webdav-scan", "http-iis-webdav-vuln"]},
    "110": {"servicio": "POP3", "scripts": ["pop3-capabilities"]},
    "111": {"servicio": "RPCBIND", "scripts": ["rpcinfo"]},
    "135": {"servicio": "RPC", "scripts": ["rpc-grind"]},
    "139": {"servicio": "NETBIOS", "scripts": ["smb-enum-shares", "smb-enum-users", "smb-os-discovery"]},
    "143": {"servicio": "IMAP", "scripts": ["imap-capabilities"]},
    "161": {"servicio": "SNMP", "scripts": ["snmp-info", "snmp-sysdescr", "snmp-processes"]},
    "389": {"servicio": "LDAP", "scripts": ["ldap-search", "ldap-rootdse"]},
    "443": {"servicio": "HTTPS", "scripts": ["ssl-cert", "ssl-enum-ciphers", "http-title", "http-headers", "http-methods", "http-robots.txt"]},
    "445": {"servicio": "SMB", "scripts": ["smb-enum-shares", "smb-enum-users", "smb-os-discovery", "smb-protocols", "smb-security-mode"]},
    "465": {"servicio": "SMTPS", "scripts": ["smtp-commands"]},
    "587": {"servicio": "SMTP", "scripts": ["smtp-commands", "smtp-open-relay"]},
    "636": {"servicio": "LDAPS", "scripts": ["ssl-cert", "ssl-enum-ciphers"]},
    "993": {"servicio": "IMAPS", "scripts": ["imap-capabilities"]},
    "995": {"servicio": "POP3S", "scripts": ["pop3-capabilities"]},
    "3306": {"servicio": "MySQL", "scripts": ["mysql-info", "mysql-users", "mysql-empty-password", "mysql-audit"]},
    "3389": {"servicio": "RDP", "scripts": ["rdp-enum-encryption"]},
    "5432": {"servicio": "PostgreSQL", "scripts": ["pgsql-brute"]},
    "5984": {"servicio": "CouchDB", "scripts": ["couchdb-databases"]},
    "6379": {"servicio": "Redis", "scripts": ["redis-info"]},
    "8080": {"servicio": "HTTP-ALT", "scripts": ["http-title", "http-headers", "http-methods", "http-robots.txt"]},
    "8443": {"servicio": "HTTPS-ALT", "scripts": ["ssl-cert", "http-title", "http-headers"]},
    "27017": {"servicio": "MongoDB", "scripts": ["mongodb-info"]},
    "50070": {"servicio": "Hadoop", "scripts": ["http-title"]}
}


def buscar_archivos_gnmap(ruta=None):
    """Busca todos los archivos .gnmap en la ruta especificada o actual"""
    if ruta is None or ruta.strip() == "":
        ruta = os.getcwd()
    
    patron = os.path.join(ruta, "*.gnmap")
    archivos = glob.glob(patron)
    
    return archivos


def buscar_archivos_xml(ruta=None):
    """Busca todos los archivos XML de Nmap en la ruta especificada o actual"""
    if ruta is None or ruta.strip() == "":
        ruta = os.getcwd()
    
    # Buscar archivos nmap*.xml recursivamente
    archivos = []
    for root, dirs, files in os.walk(ruta):
        for file in files:
            if file.startswith("nmap") and file.endswith(".xml"):
                archivos.append(os.path.join(root, file))
    
    return archivos


def buscar_archivos_nmap(ruta=None):
    """Busca todos los archivos .nmap en la ruta especificada o actual"""
    if ruta is None or ruta.strip() == "":
        ruta = os.getcwd()
    
    patron = os.path.join(ruta, "*.nmap")
    archivos = glob.glob(patron)
    
    return archivos


def extraer_datos_xml(archivo_xml):
    """Extrae información detallada del archivo XML de Nmap, incluyendo HOSTNAMES"""
    datos = []
    
    try:
        tree = ET.parse(archivo_xml)
        root = tree.getroot()
        
        # Iterar sobre cada host
        for host in root.findall(".//host"):
            ip = host.find("address[@addrtype='ipv4']")
            if ip is None:
                continue
            ip_addr = ip.get("addr")
            
            # Extraer Hostnames
            hostnames_list = []
            hostnames_elem = host.find("hostnames")
            if hostnames_elem is not None:
                for hn in hostnames_elem.findall("hostname"):
                    name = hn.get("name")
                    if name:
                        hostnames_list.append(name)
            hostnames_str = ", ".join(hostnames_list)

            # Iterar sobre cada puerto
            for port in host.findall(".//port"):
                puerto = port.get("portid")
                estado_elem = port.find("state")
                estado = estado_elem.get("state") if estado_elem is not None else "unknown"
                
                if estado != "open":
                    continue
                
                servicio_elem = port.find("service")
                servicio_info = {
                    "nombre": servicio_elem.get("name", "") if servicio_elem is not None else "",
                    "producto": servicio_elem.get("product", "") if servicio_elem is not None else "",
                    "version": servicio_elem.get("version", "") if servicio_elem is not None else "",
                    "extrainfo": servicio_elem.get("extrainfo", "") if servicio_elem is not None else "",
                    "ostype": servicio_elem.get("ostype", "") if servicio_elem is not None else "",
                    "method": servicio_elem.get("method", "") if servicio_elem is not None else "",
                    "conf": servicio_elem.get("conf", "") if servicio_elem is not None else "",
                    "cpes": []
                }
                
                # Extraer CPEs (identidades de productos)
                if servicio_elem is not None:
                    for cpe in servicio_elem.findall("cpe"):
                        servicio_info["cpes"].append(cpe.text)
                
                # Extraer información de scripts
                scripts_info = {}
                for script in port.findall("script"):
                    script_id = script.get("id", "")
                    script_output = script.get("output", "")
                    
                    # Limitar longitud de output para mejor manejo
                    if len(script_output) > 500:
                        script_output = script_output[:500] + "..."
                    
                    scripts_info[script_id] = script_output
                
                datos.append({
                    "ip": ip_addr,
                    "hostnames": hostnames_str,
                    "puerto": puerto,
                    "estado": estado,
                    "servicio": servicio_info["nombre"],
                    "producto": servicio_info["producto"],
                    "version": servicio_info["version"],
                    "extrainfo": servicio_info["extrainfo"],
                    "ostype": servicio_info["ostype"],
                    "metodo": servicio_info["method"],
                    "confianza": servicio_info["conf"],
                    "cpes": " | ".join(servicio_info["cpes"]),
                    "scripts": scripts_info,
                    "archivo_origen": os.path.basename(archivo_xml)
                })
    
    except Exception as e:
        print(f"⚠ Error procesando XML {archivo_xml}: {str(e)}")
    
    return datos


def procesar_gnmap(archivo_entrada):
    """Procesa un archivo .gnmap y extrae información de hosts, puertos ABIERTOS (SIN HOSTNAMES)"""
    resultados = []
    ips_en_alcance = []
    nombre_archivo = os.path.basename(archivo_entrada)

    with open(archivo_entrada, "r") as f:
        for linea in f:
            if linea.startswith("Host:"):
                # Capturar IP (NO extraer hostnames de .gnmap porque están incompletos)
                ip_match = re.search(r"Host:\s+(\d+\.\d+\.\d+\.\d+)", linea)
                
                if not ip_match:
                    continue
                
                ip = ip_match.group(1)

                if "Ports:" in linea:
                    puertos = linea.split("Ports:")[1].split("Ignored")[0].strip()
                    for p in puertos.split(","):
                        p = p.strip()
                        partes = p.split("/")
                        if len(partes) >= 2:
                            puerto = partes[0]
                            estado = partes[1]

                            if estado != "open":
                                continue

                            protocolo = partes[2] if len(partes) >= 3 else ""
                            servicio = partes[4] if len(partes) >= 5 else ""
                            version = ""
                            if len(partes) >= 7:
                                version = partes[6]

                            resultados.append({
                                "archivo_origen": nombre_archivo,
                                "ip": ip,
                                "hostnames": "",  # NO extraer hostnames de .gnmap
                                "puerto": puerto,
                                "protocolo": protocolo,
                                "estado": estado,
                                "servicio": servicio,
                                "version": version
                            })

                            if ip not in ips_en_alcance:
                                ips_en_alcance.append(ip)

    return resultados, ips_en_alcance


def es_dominio_valido(hostname):
    """Valida que el hostname sea un dominio completo con TLD (ej: example.com, subdomain.example.net)"""
    if not hostname:
        return False
    
    # No debe ser una IP
    if re.match(r'^\d+\.\d+\.\d+\.\d+$', hostname):
        return False
    
    # Debe tener al menos un punto (para tener TLD)
    if '.' not in hostname:
        return False
    
    # Debe terminar con un TLD válido (al menos 2 caracteres después del último punto)
    # Patrón: al menos un punto seguido de 2+ caracteres alfanuméricos
    if not re.search(r'\.[a-zA-Z0-9]{2,}$', hostname):
        return False
    
    # No debe terminar con punto (dominios incompletos como "cloudfront.")
    if hostname.endswith('.'):
        return False
    
    # Validar formato básico de dominio
    # Debe contener solo letras, números, guiones y puntos
    if not re.match(r'^[a-zA-Z0-9]([a-zA-Z0-9\-\.]*[a-zA-Z0-9])?$', hostname):
        return False
    
    return True


def procesar_nmap(archivo_entrada):
    """Procesa un archivo .nmap y extrae información de hosts, puertos ABIERTOS y DOMINIOS COMPLETOS"""
    resultados = []
    ips_en_alcance = []
    nombre_archivo = os.path.basename(archivo_entrada)
    
    ip_actual = None
    hostnames_actuales = []
    
    with open(archivo_entrada, "r", encoding='utf-8', errors='ignore') as f:
        lineas = f.readlines()
    
    i = 0
    while i < len(lineas):
        linea = lineas[i].strip()
        
        # Buscar línea "Nmap scan report for"
        # Formato: "Nmap scan report for dominio.com (IP)"
        if "Nmap scan report for" in linea:
            # Resetear para nuevo host
            ip_actual = None
            hostnames_actuales = []
            
            # Patrón principal: "Nmap scan report for dominio.com (IP)"
            # Ejemplo: "Nmap scan report for d1yehye03m313f.cloudfront.net (18.160.124.76)"
            match1 = re.search(r"Nmap scan report for\s+(.+?)\s+\((\d+\.\d+\.\d+\.\d+)\)", linea)
            if match1:
                hostname = match1.group(1).strip()
                ip_actual = match1.group(2).strip()
                
                # Validar que sea un dominio completo con TLD
                if es_dominio_valido(hostname):
                    if hostname not in hostnames_actuales:
                        hostnames_actuales.append(hostname)
                # Si hay múltiples hostnames separados por comas
                elif ',' in hostname or ' ' in hostname:
                    hostnames_separados = re.split(r'[,\s]+', hostname)
                    for hn in hostnames_separados:
                        hn_clean = hn.strip()
                        if es_dominio_valido(hn_clean) and hn_clean not in hostnames_actuales:
                            hostnames_actuales.append(hn_clean)
            else:
                # Patrón 2: "Nmap scan report for IP" (sin dominio)
                match2 = re.search(r"Nmap scan report for\s+(\d+\.\d+\.\d+\.\d+)", linea)
                if match2:
                    ip_actual = match2.group(1).strip()
                else:
                    # Patrón 3: "Nmap scan report for dominio.com" (sin IP en la misma línea)
                    match3 = re.search(r"Nmap scan report for\s+(.+?)$", linea)
                    if match3:
                        hostname = match3.group(1).strip()
                        # Buscar IP en líneas siguientes (hasta 10 líneas)
                        j = i + 1
                        while j < min(i + 10, len(lineas)) and not ip_actual:
                            ip_match = re.search(r"(\d+\.\d+\.\d+\.\d+)", lineas[j])
                            if ip_match:
                                ip_actual = ip_match.group(1).strip()
                                break
                            j += 1
                        # Validar dominio completo
                        if es_dominio_valido(hostname) and hostname not in hostnames_actuales:
                            hostnames_actuales.append(hostname)
        
        # Si tenemos una IP, buscar puertos abiertos
        if ip_actual:
            # Buscar líneas de puertos: "PORT     STATE SERVICE" seguido de líneas como "80/tcp   open  http"
            if re.match(r'^\d+/\w+\s+open\s+', linea):
                partes = linea.split()
                if len(partes) >= 2:
                    puerto_protocolo = partes[0]
                    estado = partes[1]
                    
                    if estado == "open":
                        # Separar puerto y protocolo
                        puerto_match = re.match(r'^(\d+)/(\w+)', puerto_protocolo)
                        if puerto_match:
                            puerto = puerto_match.group(1)
                            protocolo = puerto_match.group(2)
                            servicio = partes[2] if len(partes) >= 3 else ""
                            version = " ".join(partes[3:]) if len(partes) > 3 else ""
                            
                            # Limpiar version si es muy larga
                            if len(version) > 100:
                                version = version[:100] + "..."
                            
                            hostnames_str = ", ".join(hostnames_actuales) if hostnames_actuales else ""
                            
                            resultados.append({
                                "archivo_origen": nombre_archivo,
                                "ip": ip_actual,
                                "hostnames": hostnames_str,
                                "puerto": puerto,
                                "protocolo": protocolo,
                                "estado": estado,
                                "servicio": servicio,
                                "version": version
                            })
                            
                            if ip_actual not in ips_en_alcance:
                                ips_en_alcance.append(ip_actual)
        
        i += 1

    return resultados, ips_en_alcance


def deduplicar_y_combinar(resultados_gnmap, resultados_xml, resultados_nmap=None):
    """Deduplica y combina información de GNMAP, XML y NMAP, incluyendo HOSTNAMES (SOLO de .nmap y .xml)"""
    diccionario = {}
    
    # Primero agregar datos de GNMAP (SIN hostnames, se dejan vacíos)
    for item in resultados_gnmap:
        clave = f"{item['ip']}:{item['puerto']}"
        
        if clave not in diccionario:
            diccionario[clave] = {
                "archivo_origen": item.get("archivo_origen", ""),
                "ip": item.get("ip", ""),
                "hostnames": "",  # GNMAP no proporciona hostnames válidos
                "puerto": item.get("puerto", ""),
                "protocolo": item.get("protocolo", ""),
                "estado": item.get("estado", "open"),
                "servicio": item.get("servicio", ""),
                "version": item.get("version", ""),
                "producto": "",
                "extrainfo": "",
                "ostype": "",
                "metodo": "",
                "confianza": "",
                "cpes": "",
                "scripts": {}
            }
        else:
            if item.get("version") and not diccionario[clave].get("version"):
                diccionario[clave]["version"] = item["version"]
            if item.get("servicio") and len(item["servicio"]) > len(diccionario[clave].get("servicio", "")):
                diccionario[clave]["servicio"] = item["servicio"]
            # NO combinar hostnames de GNMAP (siempre vacíos)
    
    # Agregar datos de NMAP (con hostnames completos)
    if resultados_nmap:
        for item in resultados_nmap:
            clave = f"{item['ip']}:{item['puerto']}"
            
            if clave not in diccionario:
                diccionario[clave] = {
                    "archivo_origen": item.get("archivo_origen", ""),
                    "ip": item.get("ip", ""),
                    "hostnames": item.get("hostnames", ""),  # Hostnames de .nmap son válidos
                    "puerto": item.get("puerto", ""),
                    "protocolo": item.get("protocolo", ""),
                    "estado": item.get("estado", "open"),
                    "servicio": item.get("servicio", ""),
                    "version": item.get("version", ""),
                    "producto": "",
                    "extrainfo": "",
                    "ostype": "",
                    "metodo": "",
                    "confianza": "",
                    "cpes": "",
                    "scripts": {}
                }
            else:
                if item.get("version") and not diccionario[clave].get("version"):
                    diccionario[clave]["version"] = item["version"]
                if item.get("servicio") and len(item["servicio"]) > len(diccionario[clave].get("servicio", "")):
                    diccionario[clave]["servicio"] = item["servicio"]
                # Combinar hostnames de nmap (son dominios completos válidos)
                if item.get("hostnames"):
                    hostnames_existentes = diccionario[clave].get("hostnames", "").split(", ") if diccionario[clave].get("hostnames") else []
                    hostnames_nuevos = item["hostnames"].split(", ") if item["hostnames"] else []
                    todos_hostnames = list(set([h.strip() for h in hostnames_existentes + hostnames_nuevos if h.strip()]))
                    if todos_hostnames:
                        diccionario[clave]["hostnames"] = ", ".join(todos_hostnames)
    
    # Luego agregar/combinar datos de XML (con hostnames válidos)
    for item in resultados_xml:
        clave = f"{item['ip']}:{item['puerto']}"
        
        if clave not in diccionario:
            diccionario[clave] = item
        else:
            # Combinar información, priorizando datos del XML
            if item.get("producto"):
                diccionario[clave]["producto"] = item["producto"]
            if item.get("version") and not diccionario[clave].get("version"):
                diccionario[clave]["version"] = item["version"]
            if item.get("extrainfo"):
                diccionario[clave]["extrainfo"] = item["extrainfo"]
            if item.get("ostype"):
                diccionario[clave]["ostype"] = item["ostype"]
            if item.get("metodo"):
                diccionario[clave]["metodo"] = item["metodo"]
            if item.get("confianza"):
                diccionario[clave]["confianza"] = item["confianza"]
            if item.get("cpes"):
                diccionario[clave]["cpes"] = item["cpes"]
            
            # Combinar hostnames de XML (son válidos)
            if item.get("hostnames"):
                hostnames_existentes = diccionario[clave].get("hostnames", "").split(", ") if diccionario[clave].get("hostnames") else []
                hostnames_nuevos = item["hostnames"].split(", ") if item["hostnames"] else []
                todos_hostnames = list(set([h.strip() for h in hostnames_existentes + hostnames_nuevos if h.strip()]))
                if todos_hostnames:
                    diccionario[clave]["hostnames"] = ", ".join(todos_hostnames)
            
            # Combinar scripts
            if item.get("scripts"):
                diccionario[clave]["scripts"].update(item["scripts"])
    
    return list(diccionario.values())


def procesar_multiples_gnmap(archivos):
    """Procesa múltiples archivos .gnmap"""
    todos_resultados = []
    todas_ips = set()
    
    for archivo in archivos:
        print(f"📄 Procesando GNMAP: {archivo}")
        resultados, ips = procesar_gnmap(archivo)
        todos_resultados.extend(resultados)
        todas_ips.update(ips)
    
    return todos_resultados, list(todas_ips)


def procesar_multiples_xml(archivos):
    """Procesa múltiples archivos XML de Nmap"""
    todos_resultados = []
    
    for archivo in archivos:
        print(f"📄 Procesando XML: {archivo}")
        resultados = extraer_datos_xml(archivo)
        todos_resultados.extend(resultados)
    
    return todos_resultados


def procesar_multiples_nmap(archivos):
    """Procesa múltiples archivos .nmap"""
    todos_resultados = []
    todas_ips = set()
    
    for archivo in archivos:
        print(f"📄 Procesando NMAP: {archivo}")
        resultados, ips = procesar_nmap(archivo)
        todos_resultados.extend(resultados)
        todas_ips.update(ips)
    
    return todos_resultados, list(todas_ips)


def generar_identificador(ips, num_archivos):
    """Genera un identificador único"""
    if not ips:
        return f"consolidado_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"consolidado_{len(ips)}hosts_{num_archivos}scans_{fecha}"


def aplicar_formato_encabezado(ws, fila=1):
    """Aplica formato estándar a los encabezados"""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[fila]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def configurar_impresion(ws):
    """Configura la impresión de la hoja para ser profesional y legible"""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.5, bottom=0.5, header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = False
    ws.print_rows = f"1:1"


def obtener_scripts_para_puerto(puerto, servicio=""):
    """Retorna los scripts Nmap apropiados para un puerto específico"""
    if puerto in SCRIPTS_POR_PUERTO:
        return SCRIPTS_POR_PUERTO[puerto]["scripts"]
    
    scripts_genericos = ["banner", "service-version"]
    return scripts_genericos


def generar_comando_nmap_inteligente(ip, puerto, servicio="", version=""):
    """Genera comando Nmap inteligente y NO INTRUSIVO específico para el puerto"""
    scripts = obtener_scripts_para_puerto(puerto, servicio)
    scripts_str = ",".join(scripts)
    rate_params = "--min-rate 20 --max-rate 50 --max-parallelism 5"
    comando = f"nmap -sV -sC {ip} -p {puerto} --script={scripts_str} {rate_params} -oN nmap_{ip}_{puerto}.txt -oX nmap_{ip}_{puerto}.xml"
    return comando


def generar_comandos_ejecutables(ip, puertos_servicios):
    """Genera un script bash ejecutable con todos los comandos para una IP"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_content = f"""#!/bin/bash
# Script de escaneo Nmap NO INTRUSIVO para {ip}
# Generado: {timestamp}
# Parámetros: --min-rate 20 --max-rate 50 --max-parallelism 5

echo "=========================================="
echo "Iniciando escaneo para IP: {ip}"
echo "Fecha: $(date)"
echo "=========================================="
mkdir -p nmap_results_{ip}_{timestamp}
cd nmap_results_{ip}_{timestamp}
LOG_FILE="escaneo_{ip}_{timestamp}.log"
echo "Escaneo iniciado: $(date)" > $LOG_FILE
"""
    for puerto, servicio, version in puertos_servicios:
        comando = generar_comando_nmap_inteligente(ip, puerto, servicio, version)
        script_content += f"\n# Puerto {puerto} ({servicio})\n"
        script_content += f"echo \"Escaneando puerto {puerto}...\" | tee -a $LOG_FILE\n"
        script_content += f"{comando}\n"
        script_content += f"echo \"Resultado guardado: nmap_{ip}_{puerto}.{{txt,xml}}\" >> $LOG_FILE\n\n"
    
    script_content += f"""
echo "=========================================="
echo "Escaneo completado para IP: {ip}"
echo "Fecha: $(date)"
echo "=========================================="
echo "Archivos de resultado en: nmap_results_{ip}_{timestamp}/"
echo "Para ver resultados: cat escaneo_{ip}_{timestamp}.log"
"""
    return script_content


def crear_hoja_resultados(wb, resultados):
    """Crea hoja de resultados consolidados - AÑADIDA COLUMNA HOSTNAMES"""
    ws = wb.active
    ws.title = "Resultados Consolidados"
    
    # Se añade la columna "Hostnames/Dominios" después de IP
    encabezados = [
        "ID", "Archivo Origen", "IP", "Hostnames/Dominios", "Puerto", "Protocolo", "Estado Puerto",
        "Servicio", "Versión", "Vulnerabilidades Detectadas", "Severidad",
        "CVSS Score", "Riesgo Calculado", "Resumen", "Impacto", "Recomendación",
        "Estado Auditoría", "Observaciones", "Fecha Detección",
        "OS Type", "Método Detección", "Confianza", "CPEs", "Scripts Ejecutados"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    # Data Validation para Severidad (Desplazado a Col K)
    dv_severidad = DataValidation(type="list", formula1='"Crítica,Alta,Media,Baja,Revisado - No vulnerable,Pendiente"', allow_blank=True)
    dv_severidad.error = "Selecciona: Crítica, Alta, Media, Baja, Revisado - No vulnerable, Pendiente"
    dv_severidad.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_severidad)
    
    # Data Validation para Estado Auditoría (Desplazado a Col Q)
    dv_estado = DataValidation(type="list", formula1='"Pendiente,En Proceso,Finalizado,Aceptado"', allow_blank=True)
    dv_estado.error = "Selecciona: Pendiente, En Proceso, Finalizado, Aceptado"
    dv_estado.errorTitle = "Entrada inválida"
    ws.add_data_validation(dv_estado)
    
    for idx, row_data in enumerate(resultados, 2):
        scripts_ejecutados = ", ".join(row_data.get("scripts", {}).keys())[:100] if row_data.get("scripts") else ""
        
        # NOTA: Las referencias de columnas han cambiado por la inserción de Hostnames en D (4)
        # Severidad ahora es K (11), CVSS es L (12), Riesgo es M (13)
        # Formula Riesgo: Severidad (K) * CVSS (L)
        
        row = [
            idx - 1,
            row_data.get("archivo_origen", ""),
            row_data.get("ip", ""),
            row_data.get("hostnames", ""),  # Nueva columna
            row_data.get("puerto", ""),
            row_data.get("protocolo", ""),
            row_data.get("estado", "open"),
            row_data.get("servicio", ""),
            row_data.get("version", ""),
            "",  # Vulnerabilidades Detectadas
            "Pendiente",  # Severidad (K)
            "",  # CVSS Score (L)
            f"=IF(L{idx}=\"\";\"\";L{idx}*IF(K{idx}=\"Crítica\";4;IF(K{idx}=\"Alta\";3;IF(K{idx}=\"Media\";2;IF(K{idx}=\"Baja\";1;0)))))",  # Riesgo Calculado (M)
            "",  # Resumen
            "",  # Impacto
            "",  # Recomendación
            "Pendiente",  # Estado Auditoría (Q)
            "",  # Observaciones
            datetime.now().strftime("%Y-%m-%d"),
            row_data.get("ostype", ""),
            row_data.get("metodo", ""),
            row_data.get("confianza", ""),
            row_data.get("cpes", ""),
            scripts_ejecutados
        ]
        ws.append(row)
        
        dv_severidad.add(f"K{idx}") # Actualizado a K
        dv_estado.add(f"Q{idx}")    # Actualizado a Q
        
        severidad_celda = ws[f"K{idx}"]
        severidad_celda.fill = PatternFill(start_color=COLORES_SEVERIDAD["Pendiente"], 
                                           end_color=COLORES_SEVERIDAD["Pendiente"], 
                                           fill_type="solid")
    
    # Ajustar anchos de columnas (Incluyendo la nueva D)
    columnas = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
    anchos = [5, 12, 13, 25, 8, 10, 12, 12, 10, 20, 12, 10, 12, 20, 20, 20, 12, 20, 12, 12, 12, 8, 25, 20]
    
    for col, ancho in zip(columnas, anchos):
        ws.column_dimensions[col].width = ancho
    
    configurar_impresion(ws)


def crear_hoja_dashboard_vulnerabilidades(wb, resultados):
    """Crea dashboard con gráficos coloreados - ACTUALIZADO REFERENCIAS DE COLUMNAS"""
    ws = wb.create_sheet(title="Dashboard Vulnerabilidades", index=1)
    
    header_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws['A1'] = "DASHBOARD DE VULNERABILIDADES"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # Estado de Revisión
    ws['A3'] = "ESTADO DE REVISIÓN"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:B3')
    
    ws['A4'] = "Estado"
    ws['B4'] = "Cantidad"
    for cell in [ws['A4'], ws['B4']]:
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
    
    estados = ["Pendiente", "En Proceso", "Finalizado", "Aceptado"]
    max_row = len(resultados) + 100
    
    # Actualizado referencia: Estado Auditoría ahora es Q
    for idx, estado in enumerate(estados, 5):
        ws[f'A{idx}'] = estado
        ws[f'B{idx}'] = f"=COUNTIF('Resultados Consolidados'!$Q$2:$Q${max_row},A{idx})"
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    
    # Severidad
    ws['D3'] = "VULNERABILIDADES POR SEVERIDAD"
    ws['D3'].font = Font(bold=True, size=11)
    ws['D3'].fill = header_fill
    ws.merge_cells('D3:E3')
    
    ws['D4'] = "Severidad"
    ws['E4'] = "Cantidad"
    for cell in [ws['D4'], ws['E4']]:
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Actualizado referencia: Severidad ahora es K
    severidades = ["Crítica", "Alta", "Media", "Baja", "Revisado - No vulnerable", "Pendiente"]
    for idx, severidad in enumerate(severidades, 5):
        celda_sev = ws[f'D{idx}']
        celda_sev.value = severidad
        celda_sev.fill = PatternFill(start_color=COLORES_SEVERIDAD[severidad], 
                                      end_color=COLORES_SEVERIDAD[severidad], 
                                      fill_type="solid")
        celda_sev.font = Font(bold=True, color="FFFFFF" if severidad in ["Crítica", "Alta"] else "000000")
        
        ws[f'E{idx}'] = f"=COUNTIF('Resultados Consolidados'!$K$2:$K${max_row},D{idx})"
    
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    
    # Gráficos
    pie1 = PieChart()
    pie1.title = "Distribución por Estado de Revisión"
    pie1.style = 10
    labels1 = Reference(ws, min_col=1, min_row=5, max_row=8)
    data1 = Reference(ws, min_col=2, min_row=4, max_row=8)
    pie1.add_data(data1, titles_from_data=True)
    pie1.set_categories(labels1)
    ws.add_chart(pie1, "A10")
    
    pie2 = PieChart()
    pie2.title = "Distribución por Severidad"
    pie2.style = 11
    labels2 = Reference(ws, min_col=4, min_row=5, max_row=10)
    data2 = Reference(ws, min_col=5, min_row=4, max_row=10)
    pie2.add_data(data2, titles_from_data=True)
    pie2.set_categories(labels2)
    
    colores_hex = [COLORES_SEVERIDAD[s] for s in severidades]
    for idx, color_hex in enumerate(colores_hex):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color_hex
        pie2.series[0].data_points.append(pt)
    
    ws.add_chart(pie2, "D10")
    configurar_impresion(ws)


def crear_hoja_resumen_ips(wb, resultados):
    """Crea hoja de resumen por IP - ACTUALIZADO REFERENCIAS DE COLUMNAS"""
    ws = wb.create_sheet(title="Resumen por IP")
    ws.append(["IP", "Dominios Detectados", "Total Puertos", "Servicios Principales", "Riesgo Máximo", "Riesgo Promedio"])
    aplicar_formato_encabezado(ws)
    
    ips_dict = {}
    for row_data in resultados:
        ip = row_data.get("ip", "")
        if ip not in ips_dict:
            ips_dict[ip] = []
        ips_dict[ip].append(row_data)
    
    max_row_res = len(resultados) + 100
    
    for idx, (ip, datos) in enumerate(sorted(ips_dict.items()), 2):
        servicios = list(set([d.get("servicio", "Unknown") for d in datos if d.get("servicio")]))
        servicios_str = ", ".join(servicios[:5])
        
        # Obtener hostnames (usando el primero que encuentre no vacío)
        hostnames = ""
        for d in datos:
            if d.get("hostnames"):
                hostnames = d.get("hostnames")
                break

        # Referencias: IP (Col C), Riesgo Calculado (Col M)
        ws.append([
            ip,
            hostnames,
            len(datos),
            servicios_str,
            f"=IFERROR(MAX(IF('Resultados Consolidados'!$C$2:$C${max_row_res}=A{idx},'Resultados Consolidados'!$M$2:$M${max_row_res})),\"\")",
            f"=IFERROR(AVERAGE(IF('Resultados Consolidados'!$C$2:$C${max_row_res}=A{idx},'Resultados Consolidados'!$M$2:$M${max_row_res})),\"\")"
        ])
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    configurar_impresion(ws)


def crear_hoja_seguimiento_vulnerabilidades(wb, resultados):
    """Crea hoja de seguimiento centralizado"""
    ws = wb.create_sheet(title="Seguimiento Vulnerabilidades")
    
    encabezados = [
        "Hallazgo ID", "IP", "Hostname", "Puerto", "Servicio", "Vulnerabilidad",
        "Severidad", "CVSS", "Estado Remediación", "Fecha Detección",
        "Fecha Resolución", "Responsable", "Observaciones"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    dv_estado_rem = DataValidation(type="list", 
                                    formula1='"Identificada,En Remediación,Remediada,Aceptada"', 
                                    allow_blank=True)
    ws.add_data_validation(dv_estado_rem)
    
    for idx, row_data in enumerate(resultados, 2):
        ws.append([
            f"HALL-{idx-1:04d}",
            row_data.get("ip", ""),
            row_data.get("hostnames", ""),
            row_data.get("puerto", ""),
            row_data.get("servicio", ""),
            "",
            "Pendiente",
            "",
            "Identificada",
            datetime.now().strftime("%Y-%m-%d"),
            "",
            "",
            ""
        ])
        dv_estado_rem.add(f"I{idx}") # Columna I para Estado
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 15
    
    configurar_impresion(ws)


def crear_hoja_matriz_riesgos(wb, resultados):
    """Crea matriz de riesgos - ACTUALIZADO REFERENCIAS A SEGUIMIENTO"""
    ws = wb.create_sheet(title="Matriz de Riesgos")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = "MATRIZ DE RIESGOS Y GESTIÓN DE VULNERABILIDADES"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:F1')
    
    ws['A3'] = "MATRIZ: SEVERIDAD x ESTADO REMEDIACIÓN"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:E3')
    
    ws['A4'] = "Severidad"
    ws['B4'] = "Identificada"
    ws['C4'] = "En Remediación"
    ws['D4'] = "Remediada"
    ws['E4'] = "Total"
    
    for cell in [ws['A4'], ws['B4'], ws['C4'], ws['D4'], ws['E4']]:
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.font = header_font
    
    max_row_vuln = len(resultados) + 100
    severidades = ["Crítica", "Alta", "Media", "Baja"]
    
    # En Seguimiento: Severidad es Col G (7), Estado es Col I (9)
    for idx, severidad in enumerate(severidades, 5):
        celda = ws[f'A{idx}']
        celda.value = severidad
        celda.fill = PatternFill(start_color=COLORES_SEVERIDAD[severidad], 
                                 end_color=COLORES_SEVERIDAD[severidad], 
                                 fill_type="solid")
        celda.font = Font(bold=True, color="FFFFFF" if severidad in ["Crítica", "Alta"] else "000000")
        
        ws[f'B{idx}'] = f"=COUNTIFS('Seguimiento Vulnerabilidades'!$G$2:$G${max_row_vuln},A{idx},'Seguimiento Vulnerabilidades'!$I$2:$I${max_row_vuln},\"Identificada\")"
        ws[f'C{idx}'] = f"=COUNTIFS('Seguimiento Vulnerabilidades'!$G$2:$G${max_row_vuln},A{idx},'Seguimiento Vulnerabilidades'!$I$2:$I${max_row_vuln},\"En Remediación\")"
        ws[f'D{idx}'] = f"=COUNTIFS('Seguimiento Vulnerabilidades'!$G$2:$G${max_row_vuln},A{idx},'Seguimiento Vulnerabilidades'!$I$2:$I${max_row_vuln},\"Remediada\")"
        ws[f'E{idx}'] = f"=B{idx}+C{idx}+D{idx}"
    
    ws['A9'] = "TOTAL"
    ws['A9'].font = Font(bold=True)
    ws['A9'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws['B9'] = "=SUM(B5:B8)"
    ws['C9'] = "=SUM(C5:C8)"
    ws['D9'] = "=SUM(D5:D8)"
    ws['E9'] = "=SUM(E5:E8)"
    
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Matriz de Riesgos por Severidad y Estado"
    data = Reference(ws, min_col=2, min_row=4, max_col=4, max_row=8)
    cats = Reference(ws, min_col=1, min_row=5, max_row=8)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "A11")
    configurar_impresion(ws)


def crear_hoja_analisis_ip(wb, resultados):
    """Crea tabla dinámica de análisis por IP - ACTUALIZADO REFERENCIAS"""
    ws = wb.create_sheet(title="Análisis por IP")
    
    encabezados = [
        "IP", "Hostname", "Total Puertos", "Servicios Únicos", "Puertos Críticos",
        "Puertos Altos", "Puertos Medios", "Puertos Bajos",
        "Pendientes", "En Proceso", "Finalizados", "No Vulnerables"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    ips_dict = {}
    for row_data in resultados:
        ip = row_data.get("ip", "")
        if ip not in ips_dict:
            ips_dict[ip] = []
        ips_dict[ip].append(row_data)
    
    max_row_res = len(resultados) + 100
    
    for idx, (ip, datos) in enumerate(sorted(ips_dict.items()), 2):
        servicios = list(set([d.get("servicio", "") for d in datos if d.get("servicio")]))
        hostnames = ""
        for d in datos:
            if d.get("hostnames"):
                hostnames = d.get("hostnames")
                break
        
        # Referencias en Consolidados:
        # IP = C, Severidad = K, Estado = Q
        ws.append([
            ip, hostnames, len(datos), len(servicios),
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Crítica\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Alta\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Media\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Baja\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$Q$2:$Q${max_row_res},\"Pendiente\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$Q$2:$Q${max_row_res},\"En Proceso\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$Q$2:$Q${max_row_res},\"Finalizado\")",
            f"=COUNTIFS('Resultados Consolidados'!$C$2:$C${max_row_res},A{idx},'Resultados Consolidados'!$K$2:$K${max_row_res},\"Revisado - No vulnerable\")"
        ])
    
    for col in range(1, 13):
        ws.column_dimensions[chr(64 + col)].width = 14
    ws.column_dimensions['B'].width = 25
    configurar_impresion(ws)


def crear_hoja_comandos_mejorada(wb, resultados):
    """Crea hoja de comandos Nmap inteligentes"""
    ws = wb.create_sheet(title="Comandos Nmap Inteligentes")
    
    encabezados = [
        "IP", "Hostname", "Puerto", "Servicio", "Scripts Disponibles",
        "Comando Nmap (NO INTRUSIVO)", "Archivo Output", "Estado"
    ]
    
    ws.append(encabezados)
    aplicar_formato_encabezado(ws)
    
    ips_puertos = {}
    for row_data in resultados:
        ip = row_data.get("ip", "")
        hostname = row_data.get("hostnames", "")
        puerto = row_data.get("puerto", "")
        servicio = row_data.get("servicio", "")
        version = row_data.get("version", "")
        
        if ip not in ips_puertos:
            ips_puertos[ip] = []
        ips_puertos[ip].append((puerto, servicio, version, hostname))
    
    for ip, puertos_lista in sorted(ips_puertos.items()):
        for puerto, servicio, version, hostname in sorted(puertos_lista):
            scripts = obtener_scripts_para_puerto(puerto, servicio)
            scripts_str = ", ".join(scripts)
            
            comando = generar_comando_nmap_inteligente(ip, puerto, servicio, version)
            archivo_output = f"nmap_{ip}_{puerto}.txt"
            
            ws.append([
                ip, hostname, puerto, servicio, scripts_str,
                comando, archivo_output, "Pendiente"
            ])
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 100
    configurar_impresion(ws)


def crear_hoja_info_escaneos(wb, archivos_procesados):
    """Crea hoja con información de escaneos"""
    ws = wb.create_sheet(title="Info Escaneos")
    ws.append(["Archivo Procesado", "Ruta Completa", "Fecha Procesamiento"])
    aplicar_formato_encabezado(ws)
    for archivo in archivos_procesados:
        ws.append([os.path.basename(archivo), archivo, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    configurar_impresion(ws)


def crear_hoja_instrucciones(wb):
    """Crea hoja con instrucciones de uso"""
    ws = wb.create_sheet(title="Instrucciones")
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws['A1'] = "GUÍA DE USO - GESTOR DE VULNERABILIDADES v9"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')
    
    instrucciones = [
        "", "NOVEDADES v9.1:", "=================", 
        "1. SOPORTE PARA ARCHIVOS .NMAP",
        "   - Procesa archivos .nmap además de .gnmap y .xml",
        "   - Extrae dominios COMPLETOS (con TLD) de líneas 'Nmap scan report for'",
        "   - Identifica múltiples subdominios por IP",
        "   - NO usa hostnames de .gnmap (están incompletos, sin TLD)",
        "",
        "NOVEDADES v9:", "=================", 
        "1. COLUMNA HOSTNAMES/DOMINIOS",
        "   - Nueva columna en 'Resultados Consolidados' y otras hojas",
        "   - Extrae nombres DNS de archivos XML (<hostnames>)",
        "   - Extrae dominios completos de archivos NMAP (líneas de reporte)",
        "   - NO extrae hostnames de archivos GNMAP (incompletos)",
        "",
        "HOJAS DISPONIBLES:", "=================",
        "1. RESULTADOS CONSOLIDADOS",
        "   - Columna Hostnames añadida después de IP",
        "   - Riesgo Calculado ajustado a nuevas columnas",
        "",
        "2. DASHBOARD VULNERABILIDADES",
        "   - Gráficos actualizados dinámicamente",
        "",
        "3. RESUMEN POR IP",
        "   - Incluye columna de dominios asociados a la IP",
        "",
        "4. COMANDOS NMAP INTELIGENTES",
        "   - Incluye referencia del hostname",
        ""
    ]
    
    for idx, instruccion in enumerate(instrucciones, 3):
        ws[f'A{idx}'] = instruccion
    
    ws.column_dimensions['A'].width = 100
    configurar_impresion(ws)


def generar_scope_testssl(resultados, identificador, carpeta_base="resultados"):
    """Genera archivo scope para testssl.sh con formato host:puerto"""
    carpeta = f"{carpeta_base}_{identificador}"
    os.makedirs(carpeta, exist_ok=True)
    archivo_scope = os.path.join(carpeta, f"scope_testssl_{identificador}.txt")

    pares = set()
    for row in resultados:
        puerto = str(row.get("puerto", "")).strip()
        if puerto in PUERTOS_SSL_RECOMENDADOS:
            ip = row.get("ip", "").strip()
            # Si hay hostname, usarlo preferentemente para testssl si se desea, 
            # pero por consistencia usamos IP o Hostname según preferencia.
            # Aquí usamos IP:puerto por defecto, pero se podría cambiar.
            if ip:
                pares.add(f"{ip}:{puerto}")

    with open(archivo_scope, "w") as f:
        f.write("# Scope para testssl.sh - Puertos SSL/TLS detectados\n")
        f.write("# Generado automáticamente por Gestor de Vulnerabilidades\n")
        f.write(f"# Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("# Formato: host:puerto\n\n")
        for entrada in sorted(pares):
            f.write(entrada + "\n")

    print(f"✅ Scope para testssl generado: {archivo_scope}")
    return archivo_scope


def guardar_xlsx_completo(resultados, identificador, archivos_procesados, carpeta_base="resultados"):
    """Guarda resultados en XLSX con todas las hojas"""
    carpeta = f"{carpeta_base}_{identificador}"
    os.makedirs(carpeta, exist_ok=True)

    archivo_xlsx = os.path.join(carpeta, f"auditoria_{identificador}.xlsx")
    wb = openpyxl.Workbook()
    
    crear_hoja_resultados(wb, resultados)
    crear_hoja_dashboard_vulnerabilidades(wb, resultados)
    crear_hoja_resumen_ips(wb, resultados)
    crear_hoja_seguimiento_vulnerabilidades(wb, resultados)
    crear_hoja_matriz_riesgos(wb, resultados)
    crear_hoja_analisis_ip(wb, resultados)
    crear_hoja_comandos_mejorada(wb, resultados)
    crear_hoja_info_escaneos(wb, archivos_procesados)
    crear_hoja_instrucciones(wb)
    
    wb.save(archivo_xlsx)
    print(f"✅ Excel profesional generado: {archivo_xlsx}")
    
    return carpeta


def guardar_scripts_ejecutables(resultados, carpeta_scripts):
    """Genera scripts bash ejecutables por IP"""
    os.makedirs(carpeta_scripts, exist_ok=True)
    
    ips_puertos = {}
    for row_data in resultados:
        ip = row_data.get("ip", "")
        puerto = row_data.get("puerto", "")
        servicio = row_data.get("servicio", "")
        version = row_data.get("version", "")
        
        if ip not in ips_puertos:
            ips_puertos[ip] = []
        ips_puertos[ip].append((puerto, servicio, version))
    
    for ip, puertos_lista in sorted(ips_puertos.items()):
        script_content = generar_comandos_ejecutables(ip, puertos_lista)
        archivo_script = os.path.join(carpeta_scripts, f"escaneo_{ip}.sh")
        with open(archivo_script, "w") as f:
            f.write(script_content)
        os.chmod(archivo_script, 0o755)
    
    print(f"✅ Scripts de escaneo generados en: {carpeta_scripts}")


def guardar_alcance(ips, identificador, carpeta):
    """Guarda archivo de alcance"""
    archivo_salida = os.path.join(carpeta, f"alcance_{identificador}.txt")
    with open(archivo_salida, "w") as f:
        f.write(f"# IPs con puertos abiertos - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"# Total: {len(ips)} hosts\n\n")
        for ip in sorted(ips):
            f.write(ip + "\n")
    print(f"✅ Archivo de alcance generado: {archivo_salida}")


if __name__ == "__main__":
    print("=" * 80)
    print("  GESTOR PROFESIONAL DE VULNERABILIDADES - VERSIÓN 9.1 (CON DOMINIOS)")
    print("  ✓ Columna 'Hostnames/Dominios' añadida")
    print("  ✓ Captura dominios completos SOLO desde .nmap y .xml")
    print("  ✓ NO usa hostnames de .gnmap (incompletos)")
    print("=" * 80)
    print()
    
    ruta = input("Ingresa la ruta de los archivos .gnmap, .nmap y XML (Enter para carpeta actual): ").strip()
    
    archivos_gnmap = buscar_archivos_gnmap(ruta if ruta else None)
    archivos_nmap = buscar_archivos_nmap(ruta if ruta else None)
    archivos_xml = buscar_archivos_xml(ruta if ruta else None)
    
    if not archivos_gnmap and not archivos_nmap and not archivos_xml:
        print("⚠ No se encontraron archivos .gnmap, .nmap ni XML.")
        exit(1)
    
    print("\n🔄 Procesando archivos...")
    
    resultados_gnmap, ips_gnmap = procesar_multiples_gnmap(archivos_gnmap)
    resultados_nmap, ips_nmap = procesar_multiples_nmap(archivos_nmap)
    resultados_xml = procesar_multiples_xml(archivos_xml)
    
    # Combinar todas las IPs
    todas_ips = list(set(ips_gnmap + ips_nmap))
    
    resultados = deduplicar_y_combinar(resultados_gnmap, resultados_xml, resultados_nmap)
    identificador = generar_identificador(todas_ips, len(archivos_gnmap) + len(archivos_nmap) + len(archivos_xml))
    
    if resultados:
        print(f"\n📊 Resultados:")
        print(f"   - Hosts: {len(todas_ips)}")
        print(f"   - Puertos abiertos: {len(resultados)}")
        
        carpeta = f"resultados_{identificador}"
        os.makedirs(carpeta, exist_ok=True)
        
        archivos_todos = archivos_gnmap + archivos_nmap + archivos_xml
        guardar_xlsx_completo(resultados, identificador, archivos_todos)
        
        carpeta_scripts = os.path.join(carpeta, "nmap_scripts")
        guardar_scripts_ejecutables(resultados, carpeta_scripts)
        
        guardar_alcance(todas_ips, identificador, carpeta)
        generar_scope_testssl(resultados, identificador, carpeta_base="resultados")
        
        print(f"\n✅ Completado exitosamente")
        print(f"📁 Resultados: {carpeta}")
    else:
        print("⚠ No hay datos para procesar.")
